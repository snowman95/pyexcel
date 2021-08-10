#!/usr/bin/env python
#-*- coding: utf-8 -*-
"""python으로 직접 excel을 읽어오는 라이브러리
   ※ 본 문서는 Google Style Python docstring 으로 작성됨

   자료 구조
   openpyxl.Workbook == 실제 Excle 파일 1개
   - openpyxl.Workbook.Worksheet == 실제 Excle 파일 내부 시트 (pandas.DataFrame으로 상호 변환가능)

   아래의 기능을 제공한다.
   새 openpyxl.Excel.Workbook 생성
   기존 Excel 파일의 Workbook 읽어오기
   Workbook에서 지정한 WorkSheet를 Dict(sheet_name:pandas.DataFrame) 으로 반환
   DataFrame을 Excel 파일에 저장
   
Note:
   본 스크립트는 f-string을 사용하므로 Python 3.6 version 이상 사용 가능하다.
   excel을 다루는 함수는 openpyxl 라이브러리를 기반으로 한다.
   보안 프로그램으로 인해 excel 파일 로드가 불가한 경우 pywinexceltool.py를 사용하라.
   만약 문제가 생길 시 Ctrl+Shift+Esc 로 작업관리자 실행하여 Excel 프로세스를 수동으로 중단시켜라.

Example:
    import pyexceltool.py

   새 openpyxl.Excel.Workbook 생성
   workbook = create_new_workbook(sheet_name=["Sheet1", "Sheet2"])

   기존 Excel 파일의 Workbook 읽어오기
   workbook = load_workbook_with_path(load_file_path)
    
   Workbook에서 지정한 WorkSheet를 Dict(sheet_name:pandas.DataFrame) 으로 반환
   df_dict = convert_worksheet_to_df(workbook, sheet_name=["Sheet1","Sheet2"], include_index=False, include_column=False)
   DataFrame을 건드리는 작업은 pandastool.py를 가져와서 사용하도록 하자.

   DataFrame을 Excel 파일에 저장
   save_df_to_excel(df_dict, workbook, sheet_name="Sheet", file_path=save_file_path, include_index=False, include_column=False)



"""

import sys, subprocess, os, datetime
from itertools import islice
try:
    import pandas as pd
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'pandas'])
finally:
    import pandas as pd

try: import openpyxl
except ImportError: subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])
finally: import openpyxl

from openpyxl.utils.dataframe import dataframe_to_rows

### Global Value
current_time = datetime.datetime.today()
today_date = current_time.date()
default_save_directory = r'C:\\Users\\'
default_save_path = f"{default_save_directory}pyexceltool_result_{today_date}.xlsx"

def show_worksheet_list(workbook):
    """ Workbook의 sheet 목록 조회

    Args:
        workbook (openpyxl.workbook): Workbook 객체

    """
    workbook_sheet_list = workbook.sheetnames
    print(f'''workbook_sheet_list = {workbook_sheet_list}''')

def create_new_workbook(sheet_name=['Sheet']):
    """ 새로운 Excel 객체를 생성하여 반환

    Args:
        sheet_name (list): 생성할 시트명

    Returns:
        openpyxl.Workbook

    """
    workbook = openpyxl.Workbook()
    for sheet in sheet_name:
        workbook.create_sheet(title=sheet,index=None)
    return workbook

def load_workbook_with_path(file_path):
    """ 기존 Excel 파일의 Workbook을 가져옴

    Args:
        file_path (str): 불러올 파일의 경로

    Returns:
        openpyxl.Workbook

    Raises:
        ValueError: File path를 입력하지 않은 경우
        FileNotFoundError: File path의 경로에 파일 찾을 수 없는 경우
        PermissionError: File path의 파일이 이미 열려있는 경우

    """
    #1. file_path: error handling, check whether existing
    if file_path is None:
        raise ValueError("[Error] Should be given the excel file name")
    if not os.path.isabs(file_path): # in case of not full path
        file_path = os.path.join(os.getcwd(),file_path)
    if not os.path.exists(file_path):
        raise ValueError("[Error] Can't find the file from the given file path")

    # <class 'openpyxl.worksheet.worksheet.Worksheet'>
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise ValueError(r'[Error] Excel file not found. check the filepath: {file_path}')
    except PermissionError:
        raise ValueError(r'[Error] Excel file is open or the same file name exists: {file_path}')
    except Exception as err:
        raise ValueError(r"[Error] Can't open the file: {file_path}") from err
    return workbook

def convert_worksheet_to_df(workbook, sheet_name=['Sheet'], include_index=False, include_column=False):
    """ Excel 파일에서 원하는 Sheet를 지정해 list(Dataframe) 으로 반환

    Args:
        file_path (str): 저장할 대상 Excel Workbook
        sheet_name (list): Excel Workbook의 Sheet명
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    Returns:
        dict (sheet_name : pandas.DataFrame)
        
    Raises:
        ValueError: File path를 입력하지 않은 경우
        FileNotFoundError: File path의 경로에 파일 찾을 수 없는 경우

    """
    dfs = dict()
    for sheet in sheet_name:
        col = None
        idx = None
        datas = workbook[sheet].values
        if sheet in workbook and datas is not None:
            datas = list(datas)
            if include_index and include_column:
                col= datas[0][1:]
                idx = [r[0] for r in datas]
                datas = (islice(r, 1, None) for r in datas)
            elif include_index:
                idx = [r[0] for r in datas]
                datas = (islice(r, 1, None) for r in datas)
            elif include_column:
                col= datas[0]
            dfs[sheet]=pd.DataFrame(datas, index=idx, columns=col)

    if dfs is None:
        raise ValueError(f"All {sheet_name} is not exist")
    return dfs

def save_excel(workbook, file_path=None, include_index=False, include_column=False):
    """ Excel 파일을 file_path에 저장

    Args:
        workbook (openpyxl.Workbook, optional): 저장할 Excel Workbook 객체
        file_path (str, optional): 저장 위치 (C:\\Users\\...), 입력하지 않을 시 Default로 들어감.
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    """
    try :
        if file_path is None: # 파일 경로를 따로 주지 않은 경우
            file_path = default_save_path
        workbook.save(file_path)
    except FileNotFoundError:
        print('[Error] Excel file not found. check the file path')
    except PermissionError:
        print('[Error] Excel file is open or the same file name exists.')
        
def save_df_to_excel(df_dict, workbook, sheet_name=["Sheet"], file_path=None, include_index=False, include_column=False):
    """ DataFrame를 Excel sheet에 저장

    Args:
        df_dict (dict, optional): key:pandas.DataFrame
        workbook (openpyxl.Workbook, optional): 저장할 Excel Workbook 객체
        sheet_name (str, optional): 저장할 Excel Workbook sheet 이름
        file_path (str, optional): 저장 위치 (C:\\Users\\...), 입력하지 않을 시 Default로 들어감.
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    """
    for sheet in sheet_name:
        if sheet not in workbook:
            workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df_dict[sheet], index=include_index, header=include_column):
            workbook[sheet].append(r)
    save_excel(workbook,file_path,include_index,include_column)