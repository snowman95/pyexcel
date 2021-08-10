#!/usr/bin/env python
#-*- coding: utf-8 -*-
"""보안 상의 문제로 python으로 excel을 직접 읽을 수 없는 경우 pywin32를 사용하여 우회하는 방식으로 excel을 읽어오는 라이브러리
   ※ 본 문서는 Google Style Python docstring 으로 작성됨
  
   자료 구조
   Excel 는 코드 상의 Excel 객체
   Excel.Workbook 는 실제 Excle 파일 1개
   Excel.Workbook.Worksheet 는 실제 Excle 파일 내부 시트 (pandas.DataFrame으로 상호 변환가능)

   아래의 기능을 제공한다.
   새 win32.Excel 객체 생성
   새 win32.Excel.Workbook 생성
   기존 Excel 파일의 Workbook 읽어오기
   Workbook에서 지정한 WorkSheet를 Dict(sheet_name:pandas.DataFrame) 으로 반환
   DataFrame을 Excel 파일에 저장
   
Note:
   본 스크립트는 f-string을 사용하므로 Python 3.6 version 이상 사용 가능하다.
   반드시 생성한 Excel 객체는 excel.Quit()로 닫아주어야 한다.
   만약 문제가 생길 시 Ctrl+Shift+Esc 로 작업관리자 실행하여 Excel 프로세스를 수동으로 중단시켜라.

Todo:
    사전에 pywin32 라이브러리를 설치해야 한다.
    pip install pywin32
    pip install pypiwin32
    위 커맨드로 설치 실패 시 https://pypi.org/project/pywin32/#files 접속해서 자신의 python 버전에 맞는버전으로 다운로드 후
    pip install 파일명(경로포함) 입력하여 설치한다.

Example:
   import pywinexceltool.py
   새 win32.Excel 객체 생성 
   excel = create_new_excel_object()

   새 win32.Excel.Workbook 생성
   workbook = create_new_workbook(excel, ["Sheet1", "Sheet2"])

   기존 Excel 파일의 Workbook 읽어오기
   workbook = load_workbook_with_path(excel,load_file_path)
    
   Workbook에서 지정한 WorkSheet를 Dict(sheet_name:pandas.DataFrame) 으로 반환
   df_dict = convert_worksheet_to_df(excel, workbook, sheet_name=["Sheet1","Sheet2"], include_index=False, include_column=False)
   DataFrame을 건드리는 작업은 pandastool.py를 가져와서 사용하도록 하자.

   DataFrame을 Excel 파일에 저장
   save_df_to_excel(df_dict, workbook, sheet_name="Sheet", file_path=save_file_path, include_index=False, include_column=False)
    
   # 값
   # data = worksheet.Cells(1,1).Value
   # worksheet.Cells(1,1).Value="값"
   # worksheet.Range('A1:B2').Value="값"
   # worksheet.Range('A1:B2, A6:B8').Value="값"
   # worksheet.Range('B1').Interior.ColorIndex="값"
   
   # cell 폭 조정
   # worksheet.Columns(1).ColumnWidth=10
   # worksheet.Range("B:B").ColumnWidth=20
   
   # cell 높이 조정
   # worksheet.Rows(1).RowHeight=10
   # worksheet.Range("2:2").RowHeight=20



"""
import sys, subprocess, os, datetime
from itertools import islice
import win32com.client as win32

try:
    import win32com.client as win32
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'pypiwin32'])
    # 3.8버전 : https://files.pythonhosted.org/packages/0f/97/6c5a74830c5cfb8310fb291eadd95e8956fb5085673955cbcfb3f6b929c6/pywin32-301-cp38-cp38-win32.whl
    print("win32 설치 불가시 \
           https://pypi.org/project/pywin32/#files 접속해서 자신의 python 버전에 맞는걸로 다운로드\
           pip install 파일명(경로포함) 치면 설치됨.")
finally:
    import win32com.client as win32
try:
    import pandas as pd
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'pandas'])
finally:
    import pandas as pd

try: import openpyxl
except ImportError: subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])
finally: import openpyxl

from openpyxl.utils.dataframe import dataframe_to_rows

### Option Setting
pd.set_option('display.max_row', True)
pd.set_option('display.max_columns', None)  # 터미널 화면에 축약없이 전체내용 출력
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', None)
pd.set_option('display.date_yearfirst',True)
sys.stdout = open('output.txt', "w", encoding='utf8') # 한글 안깨지게 불러오기

### Global Value
current_time = datetime.datetime.today()
today_date = current_time.date()
default_save_directory = r'C:\\Users\\'
default_save_path = f"{default_save_directory}pyexceltool_result_{today_date}.xlsx"

def show_worksheet_list(workbook):
    """ Workbook의 sheet 목록 조회

    Args:
        workbook (win32.workbook): Workbook 객체

    """
    workbook_sheet_list =  [sheet.Name for sheet in workbook.Sheets]
    print(f'''workbook_sheet_list = {workbook_sheet_list}''')

def create_new_excel_object():
    """ 새로운 Excel 객체를 생성하여 반환

    Returns:
        win32com.Workbook

    """
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    return excel

def create_new_workbook(excel, sheet_name=['Sheet']):
    """ 새 Excel 파일을 생성하여 Workbook 반환

    Args:
        excel (ExcelObject): Excel Object
        sheet_name (list, optional): 지정할 Excel Workbook의 Sheet명

    Returns:
        win32com.Workbook

    """
    workbook = excel.Workbooks.Add()
    for sheet in sheet_name:
        worksheet = excel.Worksheets.Add()
        worksheet.Name=sheet
    return workbook

def load_workbook_with_path(excel, file_path):
    """ 기존 Excel 파일의 Workbook을 가져옴

    Args:
        excel (ExcelObject): Excel Object
        file_path (str): 불러올 파일의 경로

    Returns:
        win32com.Workbook

    Raises:
        ValueError: File path를 입력하지 않은 경우
        FileNotFoundError: File path의 경로에 파일 찾을 수 없는 경우
        PermissionError: File path의 파일이 이미 열려있는 경우

    """
    #1. file_path: error handling, check whether existing
    if file_path is None:
        raise ValueError("[Error] Should be given the excel file name")
    if not os.path.isabs(file_path): # in case of not full path
        file_path = os.path.join(os.getcwd(),file_path)
    if not os.path.exists(file_path):
        raise ValueError("[Error] Can't find the file from the given file path")

    # <class 'win32com.gen_py.00020813-0000-0000-C000-000000000046x0x1x9.Workbook.Workbook'>
    try:
        workbook = excel.Workbooks.Open(file_path)
    except FileNotFoundError:
        raise ValueError(r'[Error] Excel file not found. check the filepath: {file_path}')
    except PermissionError:
        raise ValueError(r'[Error] Excel file is open or the same file name exists: {file_path}')
    except Exception as err:
        raise ValueError(r"[Error] Can't open the file: {file_path}") from err
    return workbook

def convert_worksheet_to_df(excel, workbook, sheet_name=['Sheet'], include_index=False, include_column=False):
    """ Excel 파일에서 원하는 Sheet를 지정해 dict(Dataframe) 으로 반환
    
    Args:
        excel (ExcelObject): Excel Object
        workbook (win32com.Workbook): Excel Workbook
        sheet_name (list, optional): DataFrame으로 변환시킬 Excel Workbook의 Sheet명
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    Returns:
        dict (sheet_name : pandas.DataFrame)
        
    Raises:
        ValueError: Excle 에서 입력받은 모든 sheet_name이 존재하지 않는 경우 발생

    """
    dfs = dict()
    workbook_names = [sheet.Name for sheet in workbook.Sheets]
    for sheet in sheet_name:
        col,idx = (None,None)
        datas = workbook.Sheets[sheet].UsedRange.Value
        if sheet in workbook_names and datas is not None:
            datas = list(datas)
            if include_index and include_column:
                col= datas[0][1:]
                idx = [r[0] for r in datas]
                datas = (islice(r, 1, None) for r in datas)
            elif include_index:
                idx = [r[0] for r in datas]
                datas = (islice(r, 1, None) for r in datas)
            elif include_column:
                col= datas[0]
            dfs[sheet]=pd.DataFrame(datas, index=idx, columns=col)

    if dfs is None:
        workbook.Close()
        excel.Quit()
        raise ValueError(f"All {sheet_name} is not exist")
    return dfs

def save_excel(excel, workbook, file_path=None, include_index=False, include_column=False):
    """ Excel 파일을 저장
    
    Args:
        excel (ExcelObject): Excel Object
        workbook (openpyxl.Workbook, optional): 저장할 Excel Workbook
        file_path (str, optional): 저장 위치 (C:\\Users\\...), 입력하지 않을 시 Default로 들어감.
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    """
    try :
        if file_path is None: # 파일 경로를 따로 주지 않은 경우
            file_path = default_save_path
            print(f'[Warning] file_path is not entered. file is saved in default path: {default_save_path}')
        excel.SaveAs(file_path)
        workbook.Close()
        excel.Quit()
    except FileNotFoundError:
        print('[Error] Excel file not found. check the file path')
    except PermissionError:
        print('[Error] Excel file is open or the same file name exists.')

def save_df_to_excel(df_dict, excel, workbook, sheet_name=["Sheet"], file_path=None, include_index=False, include_column=False):
    """ DataFrame를 Excel 파일에 저장
    
    Args:
        df_dict (dict, optional): key:pandas.DataFrame
        excel (ExcelObject): Excel Object
        workbook (openpyxl.Workbook, optional): 저장할 Excel Workbook
        sheet_name (str, optional): 저장할 Excel Workbook sheet 이름
        file_path (str, optional): 저장 위치 (C:\\Users\\...), 입력하지 않을 시 Default로 들어감.
        include_index (bool, optional): index 포함여부
        include_column (bool, optional): column(header) 포함여부

    """
    for sheet in sheet_name:
        if sheet not in workbook:
            workbook.Worksheets.Add(title=sheet_name)
        for r in dataframe_to_rows(df_dict[sheet], index=include_index, header=include_column):
            workbook.Sheets[sheet].append(r)
    save_excel(excel, workbook,file_path,include_index,include_column)